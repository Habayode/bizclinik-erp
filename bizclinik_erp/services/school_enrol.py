"""School service — Phase 1: students and enrolment.

Enrolling a student creates a Customer (billing identity) FIRST, then the
Student record pointing at it, then an append-only StudentEnrolment row. Nothing
here posts to the GL — fees are billed later through the existing sales engine
against the student's Customer. Mutating calls require ``manage.school``.
"""
from __future__ import annotations

import io
import math
from datetime import date, datetime
from typing import Optional

import pandas as pd
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from .. import authz
from ..models import (AcademicSession, Customer, SchoolClass, Student,
                      StudentEnrolment, StudentStatus)


def _next_admission_no(session: Session) -> str:
    """Sequential STU-0001 admission number from the current Student count."""
    n = session.execute(select(func.count()).select_from(Student)).scalar_one()
    return f"STU-{n + 1:04d}"


def enrol_student(session: Session, *, first_name: str, last_name: str,
                  class_id: int, academic_session_id: int,
                  admission_no: Optional[str] = None,
                  guardian_name: Optional[str] = None,
                  guardian_phone: Optional[str] = None,
                  guardian_email: Optional[str] = None,
                  dob: Optional[date] = None,
                  gender: Optional[str] = None,
                  date_admitted: Optional[date] = None) -> Student:
    """Create the Customer (billing identity) FIRST, then the Student, then a
    StudentEnrolment row. Auto-generates the admission number when omitted."""
    authz.require_perm("manage.school")
    first = (first_name or "").strip()
    last = (last_name or "").strip()
    if not first or not last:
        raise ValueError("first_name and last_name are required.")
    if session.get(SchoolClass, class_id) is None:
        raise ValueError("Class not found.")
    if session.get(AcademicSession, academic_session_id) is None:
        raise ValueError("Academic session not found.")

    adm = (admission_no or "").strip() or _next_admission_no(session)
    if session.execute(select(Student).where(
            Student.admission_no == adm)).scalar_one_or_none():
        raise ValueError(f"Admission number {adm!r} already exists.")
    if session.execute(select(Customer).where(
            Customer.code == adm)).scalar_one_or_none():
        raise ValueError(f"A customer with code {adm!r} already exists.")

    full_name = f"{first} {last}"
    cust = Customer(code=adm, name=full_name, email=guardian_email,
                    phone=guardian_phone)
    session.add(cust); session.flush()

    student = Student(admission_no=adm, first_name=first, last_name=last,
                      dob=dob, gender=(gender or None), customer_id=cust.id,
                      current_class_id=class_id, status=StudentStatus.ACTIVE,
                      guardian_name=(guardian_name or None),
                      guardian_phone=(guardian_phone or None),
                      guardian_email=(guardian_email or None),
                      date_admitted=(date_admitted or date.today()))
    session.add(student); session.flush()

    enrol = StudentEnrolment(student_id=student.id,
                             academic_session_id=academic_session_id,
                             class_id=class_id, enrolment_status="ACTIVE")
    session.add(enrol); session.flush()
    return student


def list_students(session: Session, class_id: Optional[int] = None,
                  status: Optional[str] = None) -> list[Student]:
    q = select(Student).order_by(Student.admission_no)
    if class_id is not None:
        q = q.where(Student.current_class_id == class_id)
    if status is not None:
        q = q.where(Student.status == StudentStatus(status))
    return list(session.execute(q).scalars().all())


def withdraw_student(session: Session, student_id: int,
                     status: str = "WITHDRAWN") -> Student:
    """Mark a student off the roll and stamp their open enrolment row."""
    authz.require_perm("manage.school")
    student = session.get(Student, student_id)
    if student is None:
        raise ValueError("Student not found.")
    student.status = StudentStatus(status)
    student.status_date = date.today()
    student.is_active = False
    open_enrol = session.execute(
        select(StudentEnrolment).where(
            StudentEnrolment.student_id == student_id,
            StudentEnrolment.withdrawn_at.is_(None))
        .order_by(StudentEnrolment.enrolled_at.desc())).scalars().first()
    if open_enrol is not None:
        open_enrol.withdrawn_at = datetime.utcnow()
        open_enrol.enrolment_status = student.status.value
    session.flush()
    return student


def promote_student(session: Session, student_id: int, new_class_id: int,
                    academic_session_id: int) -> StudentEnrolment:
    """Move a student into a new class for a (new) session: update the
    denormalised current_class_id and append a fresh enrolment row."""
    authz.require_perm("manage.school")
    student = session.get(Student, student_id)
    if student is None:
        raise ValueError("Student not found.")
    if session.get(SchoolClass, new_class_id) is None:
        raise ValueError("Class not found.")
    if session.get(AcademicSession, academic_session_id) is None:
        raise ValueError("Academic session not found.")
    if session.execute(select(StudentEnrolment).where(
            StudentEnrolment.student_id == student_id,
            StudentEnrolment.academic_session_id == academic_session_id)
            ).scalar_one_or_none():
        raise ValueError("Student already has an enrolment for that session.")

    student.current_class_id = new_class_id
    enrol = StudentEnrolment(student_id=student_id,
                             academic_session_id=academic_session_id,
                             class_id=new_class_id, enrolment_status="ACTIVE")
    session.add(enrol); session.flush()
    return enrol


# --------------------------------------------------------------------------- #
# Bulk enrolment (Excel)                                                      #
# --------------------------------------------------------------------------- #

_STUDENT_COLS = ["first_name", "last_name", "class_code", "admission_no",
                 "gender", "dob", "guardian_name", "guardian_phone",
                 "guardian_email", "date_admitted"]


def student_template_bytes() -> bytes:
    """Downloadable .xlsx for bulk student enrolment."""
    instructions = [
        ["Trakit365 — Student enrolment template"], [""],
        ["One row per student. Uploading ENROLS each student: it creates the "
         "student, their class enrolment for the chosen session, and their "
         "billing record (so fees can be charged) — all in one step."],
        [""],
        ["Pick the academic session on the upload screen; it applies to every "
         "row in the file."],
        [""],
        ["Column", "Required?", "Notes"],
        ["first_name", "REQUIRED", "Student's first name."],
        ["last_name", "REQUIRED", "Student's surname."],
        ["class_code", "REQUIRED", "Must match a class in School Setup → Classes "
         "(e.g. JSS1A, PRY3)."],
        ["admission_no", "Optional", "Unique admission number; auto-generated "
         "(STU-0001…) when left blank."],
        ["gender", "Optional", "M or F."],
        ["dob", "Optional", "Date of birth, YYYY-MM-DD."],
        ["guardian_name", "Optional", "Parent/guardian (the fee payer)."],
        ["guardian_phone", "Optional", "Used for SMS fee reminders."],
        ["guardian_email", "Optional", "Used for emailed statements/reminders."],
        ["date_admitted", "Optional", "YYYY-MM-DD; defaults to today."],
        [""],
        ["Examples:"],
        ["first_name", "last_name", "class_code", "admission_no", "gender",
         "dob", "guardian_name", "guardian_phone"],
        ["Ada", "Okeke", "JSS1A", "", "F", "2013-05-10", "Mrs Okeke", "08030001111"],
        ["Emeka", "Bello", "PRY3", "OTA/2025/014", "M", "2017-02-01", "Mr Bello", "08030002222"],
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        pd.DataFrame(columns=_STUDENT_COLS).to_excel(xw, index=False, sheet_name="Students")
        pd.DataFrame(instructions).to_excel(xw, index=False, header=False,
                                            sheet_name="Instructions")
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        ws = xw.sheets["Students"]
        for i, _ in enumerate(_STUDENT_COLS, start=1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            ws.column_dimensions[get_column_letter(i)].width = 18
        ws.freeze_panes = "A2"
        instr = xw.sheets["Instructions"]
        for col, w in (("A", 22), ("B", 14), ("C", 62)):
            instr.column_dimensions[col].width = w
    return buf.getvalue()


def _clean(v) -> Optional[str]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s or None


def _parse_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _clean(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        return pd.to_datetime(s).date()
    except Exception:   # noqa: BLE001
        return None


def import_students(session: Session, df: pd.DataFrame, *,
                    academic_session_id: int) -> dict:
    """Bulk-enrol students from a filled template into the given session. Each
    good row goes through enrol_student (student + enrolment + billing record);
    bad rows are skipped with a message; duplicate admission numbers are skipped
    (idempotent). Returns {created, skipped, errors}."""
    authz.require_perm("manage.school")
    if session.get(AcademicSession, academic_session_id) is None:
        raise ValueError("Academic session not found.")
    df = df.rename(columns={c: str(c).strip().lower() for c in df.columns})
    if "first_name" not in df.columns or "class_code" not in df.columns:
        raise ValueError("The file needs at least 'first_name' and 'class_code' "
                         "columns — use the template.")
    classes = {c.class_code: c.id for c in
               session.execute(select(SchoolClass)).scalars()}
    existing = {a for (a,) in session.execute(select(Student.admission_no)).all()}
    seen: set = set()
    created = skipped = 0
    errors: list[str] = []

    for idx, row in df.iterrows():
        rno = int(idx) + 2
        fn, ln = _clean(row.get("first_name")), _clean(row.get("last_name"))
        cc = _clean(row.get("class_code"))
        if not fn and not ln and not cc:
            continue   # wholly blank row
        if not fn or not ln:
            errors.append(f"Row {rno}: first_name and last_name are required.")
            skipped += 1; continue
        if not cc:
            errors.append(f"Row {rno}: class_code is required.")
            skipped += 1; continue
        cid = classes.get(cc)
        if cid is None:
            errors.append(f"Row {rno}: class '{cc}' not found — add it in "
                          "School Setup → Classes.")
            skipped += 1; continue
        adm = _clean(row.get("admission_no"))
        if adm and (adm in existing or adm in seen):
            skipped += 1; continue   # already enrolled — idempotent
        try:
            with session.begin_nested():
                stu = enrol_student(
                    session, first_name=fn, last_name=ln, class_id=cid,
                    academic_session_id=academic_session_id, admission_no=adm,
                    gender=_clean(row.get("gender")),
                    dob=_parse_date(row.get("dob")),
                    guardian_name=_clean(row.get("guardian_name")),
                    guardian_phone=_clean(row.get("guardian_phone")),
                    guardian_email=_clean(row.get("guardian_email")),
                    date_admitted=_parse_date(row.get("date_admitted")))
            existing.add(stu.admission_no); seen.add(stu.admission_no)
            created += 1
        except ValueError as e:
            errors.append(f"Row {rno}: {e}")
            skipped += 1
    return {"created": created, "skipped": skipped, "errors": errors}
