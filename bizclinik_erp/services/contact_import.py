"""Bulk import of customers / suppliers from a filled Excel template.

Two pieces:
  * ``template_bytes(kind)`` -> a downloadable .xlsx with the right column
    headers (data sheet is intentionally empty so nothing is imported by
    accident) plus an Instructions sheet with a worked example.
  * ``import_rows(session, kind, df)`` -> validates a DataFrame read from the
    uploaded file and inserts the rows, auto-generating a code when blank and
    skipping duplicates. Returns a {created, skipped, errors} summary.

Used by the Settings page so a business can migrate its existing customer /
supplier list in one upload instead of typing them one by one.
"""
from __future__ import annotations

import io
import math
from typing import Optional

import pandas as pd
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..models import Customer, Supplier

# (model, columns, code-prefix, sheet-name)
_SPEC = {
    "customer": (Customer,
                 ["code", "name", "email", "phone", "address", "credit_limit"],
                 "C", "Customers"),
    "supplier": (Supplier,
                 ["code", "name", "email", "phone", "address"],
                 "S", "Suppliers"),
}

_INSTRUCTIONS = {
    "customer": [
        ["Trakit365 — Customer import template"],
        [""],
        ["Fill ONE row per customer on the 'Customers' sheet, then upload it on"],
        ["Settings -> Customers -> Bulk import."],
        [""],
        ["Column", "Required?", "Notes"],
        ["code", "Optional", "Your own customer code. Leave blank to auto-generate (C0001…)."],
        ["name", "REQUIRED", "Customer / business name."],
        ["email", "Optional", "Contact email."],
        ["phone", "Optional", "Contact phone."],
        ["address", "Optional", "Postal / delivery address."],
        ["credit_limit", "Optional", "Number, e.g. 500000. Leave blank for 0."],
        [""],
        ["Example row:"],
        ["C001", "Sunrise Restaurant Ltd", "pay@sunrise.ng", "08030000001",
         "5 Allen Ave, Ikeja", "500000"],
        [""],
        ["Notes: blank rows are ignored. A code that already exists is skipped"],
        ["(so re-uploading the same file won't create duplicates)."],
    ],
    "supplier": [
        ["Trakit365 — Supplier import template"],
        [""],
        ["Fill ONE row per supplier on the 'Suppliers' sheet, then upload it on"],
        ["Settings -> Suppliers -> Bulk import."],
        [""],
        ["Column", "Required?", "Notes"],
        ["code", "Optional", "Your own supplier code. Leave blank to auto-generate (S0001…)."],
        ["name", "REQUIRED", "Supplier / vendor name."],
        ["email", "Optional", "Contact email."],
        ["phone", "Optional", "Contact phone."],
        ["address", "Optional", "Address."],
        [""],
        ["Example row:"],
        ["S001", "FreshFarm Produce Ltd", "sales@freshfarm.ng", "07010001111",
         "Mile 12 Market, Lagos"],
        [""],
        ["Notes: blank rows are ignored. A code that already exists is skipped"],
        ["(so re-uploading the same file won't create duplicates)."],
    ],
}


def template_bytes(kind: str) -> bytes:
    """Return a styled, ready-to-fill .xlsx template for ``customer``/``supplier``."""
    model, cols, _prefix, sheet = _SPEC[kind]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        # Data sheet: headers only — nothing gets imported by accident.
        pd.DataFrame(columns=cols).to_excel(xw, index=False, sheet_name=sheet)
        pd.DataFrame(_INSTRUCTIONS[kind]).to_excel(
            xw, index=False, header=False, sheet_name="Instructions")

        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        ws = xw.sheets[sheet]
        widths = {"code": 12, "name": 34, "email": 28, "phone": 18,
                  "address": 42, "credit_limit": 14}
        for i, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)
        ws.freeze_panes = "A2"
        instr = xw.sheets["Instructions"]
        for col in ("A", "B", "C"):
            instr.column_dimensions[col].width = {"A": 24, "B": 12, "C": 60}[col]
    return buf.getvalue()


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float:
    try:
        if v is None or (isinstance(v, float) and math.isnan(v)):
            return 0.0
        return float(str(v).replace(",", "").strip() or 0)
    except (ValueError, TypeError):
        return 0.0


def _next_code(session: Session, model, prefix: str, existing: set) -> str:
    n = (session.execute(select(func.count(model.id))).scalar() or 0) + 1
    cand = f"{prefix}{n:04d}"
    while cand in existing:
        n += 1
        cand = f"{prefix}{n:04d}"
    return cand


def import_rows(session: Session, kind: str, df: pd.DataFrame) -> dict:
    """Insert rows from ``df`` (read from the uploaded template). Auto-generates
    a code when blank; skips rows with a duplicate code. Returns a summary."""
    if kind not in _SPEC:
        raise ValueError(f"Unknown import kind {kind!r}.")
    model, _cols, prefix, _sheet = _SPEC[kind]
    df = df.rename(columns={c: str(c).strip().lower() for c in df.columns})
    if "name" not in df.columns:
        raise ValueError("The file has no 'name' column — use the provided template.")

    existing = {c for (c,) in session.execute(select(model.code)).all()}
    created, skipped = 0, 0
    errors: list[str] = []

    for idx, row in df.iterrows():
        name = _clean(row.get("name"))
        if not name:
            continue                       # blank row — ignore silently
        code = _clean(row.get("code"))
        if not code:
            code = _next_code(session, model, prefix, existing)
        if code in existing:
            skipped += 1
            errors.append(f"Row {int(idx) + 2}: code '{code}' already exists — skipped.")
            continue
        kwargs = dict(code=code, name=name,
                      email=_clean(row.get("email")),
                      phone=_clean(row.get("phone")),
                      address=_clean(row.get("address")))
        if kind == "customer":
            kwargs["credit_limit"] = _num(row.get("credit_limit"))
        session.add(model(**kwargs))
        session.flush()
        existing.add(code)
        created += 1

    return {"created": created, "skipped": skipped, "errors": errors}
