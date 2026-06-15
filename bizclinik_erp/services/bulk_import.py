"""Generic bulk import of master data from a filled Excel template.

One registry (`SPECS`) drives downloadable templates + validated imports for
customers, suppliers, products, employees and chart-of-accounts. Each spec lists
its columns and how to parse them; the engine builds a styled .xlsx template
(empty data sheet + an Instructions sheet) and inserts rows, auto-generating a
code when blank (where allowed) and skipping duplicates.

Public API (unchanged from the earlier contact_import):
    template_bytes(kind) -> bytes        # downloadable .xlsx
    import_rows(session, kind, df) -> {"created", "skipped", "errors"}
    KINDS                                # available kinds + display labels
"""
from __future__ import annotations

import io
import math
from dataclasses import dataclass, field
from typing import Callable, Optional

import pandas as pd
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from ..models import (Account, AccountType, BankAccount, Customer, Employee,
                      Product, Supplier)


# --------------------------------------------------------------------------- #
# Spec model                                                                  #
# --------------------------------------------------------------------------- #

@dataclass
class Field:
    col: str                       # column header in the spreadsheet
    kind: str = "str"              # str | num | bool | enum | parent
    required: bool = False
    kwarg: Optional[str] = None    # model attr (defaults to col)
    default: object = None
    help: str = ""

    @property
    def key(self) -> str:
        return self.kwarg or self.col


@dataclass
class Spec:
    model: type
    sheet: str
    title: str
    noun: str
    code_col: str = "code"
    code_prefix: Optional[str] = None   # None => code required (no auto-gen)
    code_help: str = ""
    fields: list = field(default_factory=list)   # excludes the code col; 'name' first
    enum_type: Optional[type] = None
    example: Optional[list] = None
    # Optional hook(session, kwargs, row) -> kwargs, for special resolution
    # (e.g. linking a bank to its GL account). May raise ValueError per row.
    finalize: Optional[Callable] = None


SPECS: dict[str, Spec] = {
    "customer": Spec(
        Customer, "Customers", "Customer", "customer", code_prefix="C",
        code_help="Your own customer code. Blank = auto-generate (C0001…).",
        fields=[
            Field("name", required=True, help="Customer / business name."),
            Field("email", help="Contact email."),
            Field("phone", help="Contact phone."),
            Field("address", help="Postal / delivery address."),
            Field("credit_limit", "num", default=0.0,
                  help="Number, e.g. 500000. Blank = 0."),
        ],
        example=["C001", "Sunrise Restaurant Ltd", "pay@sunrise.ng",
                 "08030000001", "5 Allen Ave, Ikeja", "500000"]),

    "supplier": Spec(
        Supplier, "Suppliers", "Supplier", "supplier", code_prefix="S",
        code_help="Your own supplier code. Blank = auto-generate (S0001…).",
        fields=[
            Field("name", required=True, help="Supplier / vendor name."),
            Field("email", help="Contact email."),
            Field("phone", help="Contact phone."),
            Field("address", help="Address."),
        ],
        example=["S001", "FreshFarm Produce Ltd", "sales@freshfarm.ng",
                 "07010001111", "Mile 12 Market, Lagos"]),

    "product": Spec(
        Product, "Products", "Product", "product", code_col="sku",
        code_prefix="P",
        code_help="Stock-keeping unit (SKU). Blank = auto-generate (P0001…).",
        fields=[
            Field("name", required=True, help="Product / item name."),
            Field("description", help="Optional description."),
            Field("unit", default="ea", help="Unit of measure, e.g. ea, bag, kg. Blank = ea."),
            Field("standard_price", "num", default=0.0, help="Default selling price (₦)."),
            Field("standard_cost", "num", default=0.0, help="Default cost (₦)."),
            Field("reorder_level", "num", default=0.0, help="Low-stock reorder point."),
            Field("is_stockable", "bool", default=True,
                  help="yes/no — does it carry inventory? Blank = yes."),
        ],
        example=["RICE50", "Rice 50kg Bag", "Long-grain rice", "bag",
                 "45000", "38000", "10", "yes"]),

    "employee": Spec(
        Employee, "Employees", "Employee", "employee", code_prefix="EMP",
        code_help="Staff code. Blank = auto-generate (EMP0001…).",
        fields=[
            Field("name", required=True, help="Full name."),
            Field("email", help="Email."),
            Field("phone", help="Phone."),
            Field("department", help="Department, e.g. Sales."),
            Field("job_title", help="Job title."),
            Field("employment_type", help="full-time / part-time / contract / intern."),
            Field("monthly_gross", "num", default=0.0, help="Gross monthly pay (₦)."),
            Field("paye_rate", "num", default=0.0,
                  help="Flat PAYE override (0 = graduated, recommended)."),
            Field("pension_rate", "num", default=0.08, help="Employee pension rate. Blank = 0.08."),
            Field("pension_employer_rate", "num", default=0.10,
                  help="Employer pension rate. Blank = 0.10."),
            Field("annual_leave_days", "num", default=20.0,
                  help="Annual leave entitlement. Blank = 20."),
        ],
        example=["EMP001", "Chioma Okeke", "chioma@x.ng", "08030000010",
                 "Operations", "Store Manager", "full-time", "250000", "0",
                 "0.08", "0.10", "20"]),

    "account": Spec(
        Account, "Accounts", "Chart-of-accounts", "account",
        code_prefix=None,   # account code is meaningful — required, never auto
        code_help="GL account code, e.g. 4200. REQUIRED — accounts are not auto-numbered.",
        enum_type=AccountType,
        fields=[
            Field("name", required=True, help="Account name."),
            Field("type", "enum", required=True,
                  help="ASSET / LIABILITY / EQUITY / INCOME / EXPENSE."),
            Field("parent_code", "parent", kwarg="parent_id",
                  help="Optional parent account's code (must already exist)."),
            Field("is_postable", "bool", default=True,
                  help="yes/no — can entries post directly to it? Blank = yes."),
        ],
        example=["4200", "Service Revenue", "INCOME", "4000", "yes"]),

    "bank": Spec(
        BankAccount, "Bank accounts", "Bank account", "bank account",
        code_prefix="BANK",
        code_help="Your code for this account. Blank = auto-generate (BANK0001…).",
        fields=[
            Field("name", required=True, help="Account label, e.g. 'GTBank — Operating'."),
            Field("bank", help="Bank name, e.g. GTBank."),
            Field("account_number", help="Account number (kept as text)."),
            Field("gl_account_code", "ext",
                  help="Existing asset GL account code to post cash to. Blank = "
                       "a new GL account is created automatically under 1120 Bank."),
        ],
        example=["BANK1", "GTBank — Operating", "GTBank", "0123456789", "1120"]),
}

KINDS = {k: s.sheet for k, s in SPECS.items()}


def _create_bank_gl(session: Session, label: str) -> int:
    """Auto-create an asset GL account for a bank, under 1120 'Bank' if present."""
    parent_id = session.execute(
        select(Account.id).where(Account.code == "1120")).scalar()
    existing = {c for (c,) in session.execute(select(Account.code)).all()}
    code = next((str(n) for n in range(1121, 1200) if str(n) not in existing), None)
    if code is None:
        raise ValueError(
            "Ran out of auto-numbered bank GL codes (1121–1199). Put an existing "
            "asset account code in the 'gl_account_code' column instead.")
    acct = Account(code=code, name=f"Bank — {label}"[:255],
                   type=AccountType.ASSET, parent_id=parent_id, is_postable=True)
    session.add(acct)
    session.flush()
    return acct.id


def _bank_finalize(session: Session, kwargs: dict, row) -> dict:
    code = _clean(row.get("gl_account_code"))
    if code:
        aid = session.execute(
            select(Account.id).where(Account.code == code)).scalar()
        if aid is None:
            raise ValueError(
                f"GL account '{code}' not found — create it first, or leave "
                "the column blank to auto-create one.")
        kwargs["gl_account_id"] = aid
    else:
        kwargs["gl_account_id"] = _create_bank_gl(
            session, kwargs.get("name") or kwargs.get("code") or "Bank")
    return kwargs


SPECS["bank"].finalize = _bank_finalize


# --------------------------------------------------------------------------- #
# Template                                                                    #
# --------------------------------------------------------------------------- #

def _instructions(spec: Spec) -> list:
    rows = [[f"Trakit365 — {spec.title} import template"], [""],
            [f"Fill ONE row per {spec.noun} on the '{spec.sheet}' sheet, then upload it."],
            [""], ["Column", "Required?", "Notes"],
            [spec.code_col, "REQUIRED" if spec.code_prefix is None else "Optional",
             spec.code_help]]
    for f in spec.fields:
        rows.append([f.col, "REQUIRED" if f.required else "Optional", f.help])
    if spec.example:
        rows += [[""], ["Example row (the values, in column order):"], spec.example]
    rows += [[""],
             ["Notes: blank rows are ignored. A code that already exists is skipped,"],
             ["so re-uploading the same file will not create duplicates."]]
    return rows


def template_bytes(kind: str) -> bytes:
    spec = SPECS[kind]
    headers = [spec.code_col] + [f.col for f in spec.fields]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        pd.DataFrame(columns=headers).to_excel(xw, index=False, sheet_name=spec.sheet)
        pd.DataFrame(_instructions(spec)).to_excel(
            xw, index=False, header=False, sheet_name="Instructions")

        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        ws = xw.sheets[spec.sheet]
        wide = {"name", "address", "description"}
        for i, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F3864")
            ws.column_dimensions[get_column_letter(i)].width = 38 if h in wide else 16
        ws.freeze_panes = "A2"
        instr = xw.sheets["Instructions"]
        for col, w in (("A", 24), ("B", 12), ("C", 64)):
            instr.column_dimensions[col].width = w
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Import                                                                      #
# --------------------------------------------------------------------------- #

def _clean(v) -> Optional[str]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s or None


def _num(v, default: float = 0.0) -> float:
    s = _clean(v)
    if s is None:
        return default
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return default


_TRUE = {"1", "true", "t", "yes", "y", "x", "stockable", "active"}
_FALSE = {"0", "false", "f", "no", "n"}


def _bool(v, default: bool) -> bool:
    s = _clean(v)
    if s is None:
        return default
    s = s.lower()
    return True if s in _TRUE else (False if s in _FALSE else default)


def _next_code(session: Session, model, code_col: str, prefix: str,
               existing: set) -> str:
    n = (session.execute(select(func.count(model.id))).scalar() or 0) + 1
    cand = f"{prefix}{n:04d}"
    while cand in existing:
        n += 1
        cand = f"{prefix}{n:04d}"
    return cand


def import_rows(session: Session, kind: str, df: pd.DataFrame) -> dict:
    if kind not in SPECS:
        raise ValueError(f"Unknown import kind {kind!r}.")
    spec = SPECS[kind]
    df = df.rename(columns={c: str(c).strip().lower() for c in df.columns})
    if "name" not in df.columns:
        raise ValueError("The file has no 'name' column — use the provided template.")

    code_col = spec.code_col
    existing = {c for (c,) in
                session.execute(select(getattr(spec.model, code_col))).all()}
    created, skipped = 0, 0
    errors: list[str] = []

    for idx, row in df.iterrows():
        rno = int(idx) + 2          # +1 header, +1 to 1-base
        name = _clean(row.get("name"))
        if not name:
            continue
        code = _clean(row.get(code_col))
        if not code:
            if spec.code_prefix is None:
                skipped += 1
                errors.append(f"Row {rno}: '{code_col}' is required — skipped.")
                continue
            code = _next_code(session, spec.model, code_col, spec.code_prefix, existing)
        if code in existing:
            skipped += 1
            errors.append(f"Row {rno}: {code_col} '{code}' already exists — skipped.")
            continue

        kwargs = {code_col: code}
        bad = False
        for f in spec.fields:
            raw = row.get(f.col)
            if f.kind == "str":
                kwargs[f.key] = _clean(raw) if _clean(raw) is not None else f.default
            elif f.kind == "num":
                kwargs[f.key] = _num(raw, f.default if f.default is not None else 0.0)
            elif f.kind == "bool":
                kwargs[f.key] = _bool(raw, bool(f.default))
            elif f.kind == "enum":
                v = _clean(raw)
                if not v:
                    errors.append(f"Row {rno}: '{f.col}' is required."); bad = True; break
                try:
                    kwargs[f.key] = spec.enum_type(v.upper())
                except ValueError:
                    opts = ", ".join(e.value for e in spec.enum_type)
                    errors.append(f"Row {rno}: '{f.col}'='{v}' invalid (use {opts}).")
                    bad = True; break
            elif f.kind == "parent":
                pc = _clean(raw)
                pid = None
                if pc:
                    pid = session.execute(
                        select(spec.model.id).where(spec.model.code == pc)
                    ).scalar()
                    if pid is None:
                        errors.append(f"Row {rno}: parent '{pc}' not found — set to none.")
                kwargs[f.key] = pid
            elif f.kind == "ext":
                pass    # handled by spec.finalize below
        if not bad and spec.finalize is not None:
            try:
                kwargs = spec.finalize(session, kwargs, row)
            except ValueError as e:
                errors.append(f"Row {rno}: {e}")
                bad = True
        if bad:
            skipped += 1
            continue

        session.add(spec.model(**kwargs))
        session.flush()
        existing.add(code)
        created += 1

    return {"created": created, "skipped": skipped, "errors": errors}
