"""Opening-balance import — load a closing trial balance and post it as one
balanced opening journal, so a migrating business goes live with real figures.

Each row is a GL account code with a debit OR a credit. The engine resolves the
accounts, checks the trial balance balances (or absorbs a rounding/equity
difference into a chosen balancing account), and posts a single
``source_kind="OPENING"`` journal via the standard ledger chokepoint (which
re-enforces DR = CR and period rules). It refuses to run twice — void the
existing opening journal first to re-import.

Note: this loads the GL control balances (AR/AP totals, cash, stock, equity…).
Per-customer/supplier invoice-level detail comes from the (future) invoice/bill
import; the aggregate AR/AP here ties to those control accounts.
"""
from __future__ import annotations

import io
import math
from datetime import date
from typing import Optional

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import Account, JournalEntry
from .ledger import JELine, post_journal

SHEET = "Trial Balance"
SOURCE_KIND = "OPENING"
DEFAULT_PLUG_CODE = "3200"   # Retained Earnings / Opening Balance Equity


def _clean(v) -> Optional[str]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float:
    s = _clean(v)
    if s is None:
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return 0.0


def template_bytes() -> bytes:
    headers = ["account_code", "account_name", "debit", "credit"]
    instructions = [
        ["Trakit365 — Opening balances (trial balance) template"], [""],
        ["Enter your CLOSING trial balance as at the day before you go live."],
        ["One GL account per row. Put the amount in EITHER the debit OR the "
         "credit column (not both)."],
        ["'account_code' must match a code in your Chart of Accounts "
         "(General Ledger → Chart of Accounts). 'account_name' is for your "
         "reference only."],
        [""],
        ["Column", "Required?", "Notes"],
        ["account_code", "REQUIRED", "Existing GL account code, e.g. 1120."],
        ["account_name", "Optional", "Ignored on import — just a label for you."],
        ["debit", "One of", "Debit amount (₦). Assets & expenses are usually debits."],
        ["credit", "debit/credit", "Credit amount (₦). Liabilities, equity, income are usually credits."],
        [""],
        ["The debit and credit TOTALS must be equal. If they differ by a small "
         "amount, choose a balancing account on the upload screen and the "
         "difference is posted there (usually Retained Earnings / Opening "
         "Balance Equity)."],
        [""],
        ["Example rows:"],
        ["1120", "Bank — Operating", "2500000", ""],
        ["1130", "Accounts Receivable", "850000", ""],
        ["1140", "Inventory", "1200000", ""],
        ["2110", "Accounts Payable", "", "1400000"],
        ["3100", "Share Capital", "", "3150000"],
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        pd.DataFrame(columns=headers).to_excel(xw, index=False, sheet_name=SHEET)
        pd.DataFrame(instructions).to_excel(
            xw, index=False, header=False, sheet_name="Instructions")
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        ws = xw.sheets[SHEET]
        widths = {"account_code": 16, "account_name": 34, "debit": 16, "credit": 16}
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            ws.column_dimensions[get_column_letter(i)].width = widths[h]
        ws.freeze_panes = "A2"
        instr = xw.sheets["Instructions"]
        for col, w in (("A", 18), ("B", 14), ("C", 70)):
            instr.column_dimensions[col].width = w
    return buf.getvalue()


def existing_opening(session: Session) -> Optional[JournalEntry]:
    return session.execute(
        select(JournalEntry).where(JournalEntry.source_kind == SOURCE_KIND)
        .order_by(JournalEntry.id.desc())
    ).scalars().first()


def import_trial_balance(session: Session, df: pd.DataFrame, *, as_of: date,
                         plug_account_code: Optional[str] = None,
                         user_id: Optional[int] = None) -> dict:
    """Validate a trial balance and post it as one opening journal. All-or-
    nothing: raises ValueError (without posting) on any problem so the caller
    can show it and let the user fix the sheet."""
    existing = existing_opening(session)
    if existing is not None:
        raise ValueError(
            f"Opening balances were already posted ({existing.entry_no}). Void "
            "that journal first if you need to re-import.")

    df = df.rename(columns={c: str(c).strip().lower() for c in df.columns})
    if "account_code" not in df.columns:
        raise ValueError("The file has no 'account_code' column — use the template.")

    lines: list[JELine] = []
    total_dr = total_cr = 0.0
    errors: list[str] = []
    for idx, row in df.iterrows():
        rno = int(idx) + 2
        code = _clean(row.get("account_code"))
        # Round at the source so the totals (and the plug derived from them)
        # match the rounded line amounts post_journal re-sums and re-checks.
        dr = round(_num(row.get("debit")), 2)
        cr = round(_num(row.get("credit")), 2)
        if not code and dr == 0 and cr == 0:
            continue
        if not code:
            errors.append(f"Row {rno}: missing account_code.")
            continue
        acct = session.execute(
            select(Account).where(Account.code == code)).scalar_one_or_none()
        if acct is None:
            errors.append(f"Row {rno}: account '{code}' not found.")
            continue
        if not acct.is_postable:
            errors.append(f"Row {rno}: account '{code}' is a header (not postable).")
            continue
        if dr > 0 and cr > 0:
            errors.append(f"Row {rno}: account '{code}' has BOTH a debit and a credit.")
            continue
        if dr == 0 and cr == 0:
            continue
        lines.append(JELine(account_id=acct.id, debit=dr, credit=cr,
                            memo=f"Opening balance — {acct.name}"))
        total_dr += dr
        total_cr += cr

    if errors:
        raise ValueError("Fix these before importing:\n• " + "\n• ".join(errors))
    if not lines:
        raise ValueError("No opening-balance rows found in the file.")

    diff = round(total_dr - total_cr, 2)
    plug_amount = 0.0
    if abs(diff) > 0.01:
        if not plug_account_code:
            raise ValueError(
                f"Trial balance is out of balance by ₦{abs(diff):,.2f} "
                f"(debits ₦{total_dr:,.2f} vs credits ₦{total_cr:,.2f}). Fix the "
                "figures, or pick a balancing account to absorb the difference.")
        plug = session.execute(
            select(Account).where(Account.code == plug_account_code)
        ).scalar_one_or_none()
        if plug is None or not plug.is_postable:
            raise ValueError(f"Balancing account '{plug_account_code}' not found "
                             "or not postable.")
        # diff > 0 => debits exceed credits => plug a credit (and vice-versa).
        if diff > 0:
            lines.append(JELine(account_id=plug.id, credit=abs(diff),
                                memo="Opening balance — balancing entry"))
        else:
            lines.append(JELine(account_id=plug.id, debit=abs(diff),
                                memo="Opening balance — balancing entry"))
        plug_amount = abs(diff)

    je = post_journal(session, as_of, "Opening balances", lines,
                      source_kind=SOURCE_KIND, user_id=user_id)
    return {"je_no": je.entry_no, "lines": len(lines),
            "total_debit": round(total_dr, 2), "total_credit": round(total_cr, 2),
            "balancing_amount": plug_amount,
            "balancing_account": plug_account_code if plug_amount else None}
