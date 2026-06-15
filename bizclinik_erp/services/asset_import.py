"""Fixed-asset register import — migrate an existing fixed-asset register in
from an Excel template, or bulk-add new assets.

One row per asset. The importer creates the sub-ledger asset records ONLY; it
posts no journal entry — exactly like ``services.assets.add_asset``. The asset
cost and any accumulated depreciation are expected to already sit in the GL
(via the opening-balance trial balance, or the original purchase bill), so this
register import is the detail behind those control balances and never moves the
trial balance.

Migration vs. new:
  * A brand-new asset has ``accumulated_depreciation`` = 0; depreciation runs
    forward from its acquired date.
  * A migrated, already-in-service asset carries accumulated depreciation to
    date AND a "depreciation booked through" date. Both are required together
    so ``run_depreciation`` resumes from the next month instead of back-posting
    every month since acquisition.

GL accounts per row are resolved from the category (Equipment -> 1210,
Furniture -> 1220), reused by name, or auto-created under 1200 Fixed Assets for
any other category; accumulated depreciation defaults to 1290 and depreciation
expense to 6600. Every account is overridable with an explicit code column.
"""
from __future__ import annotations

import calendar
import io
import math
from datetime import date, datetime
from typing import Optional

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import authz
from ..models import (Account, AccountType, AssetStatus, DepreciationMethod,
                      FixedAsset)

SHEET = "Fixed Assets"
ACCUM_DEP_CODE = "1290"      # Accumulated Depreciation (seeded)
DEP_EXPENSE_CODE = "6600"    # Depreciation Expense (seeded)
FIXED_ASSET_PARENT = "1200"  # Fixed Assets header — parent for auto-created accounts

# Category -> default asset GL code (seeded postable accounts).
_CATEGORY_GL = {
    "equipment": "1210",
    "furniture": "1220",
    "furniture & fittings": "1220",
}


# --------------------------------------------------------------------------- #
# Parsing helpers                                                             #
# --------------------------------------------------------------------------- #

def _clean(v) -> Optional[str]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float:
    s = _clean(v)
    if s is None:
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return float("nan")


def _int(v) -> Optional[int]:
    s = _clean(v)
    if s is None:
        return None
    try:
        return int(round(float(s.replace(",", ""))))
    except ValueError:
        return None


def _parse_date(v) -> Optional[date]:
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:                                   # last resort: let pandas try
        return pd.to_datetime(s).date()
    except Exception:                      # noqa: BLE001
        return None


def _end_of_month(d: date) -> date:
    return date(d.year, d.month, calendar.monthrange(d.year, d.month)[1])


# --------------------------------------------------------------------------- #
# GL account resolution                                                       #
# --------------------------------------------------------------------------- #

def _postable_by_code(session: Session, code: str) -> Optional[Account]:
    return session.execute(
        select(Account).where(Account.code == code)).scalar_one_or_none()


def _create_asset_account(session: Session, name: str,
                          code: Optional[str] = None) -> int:
    parent_id = session.execute(
        select(Account.id).where(Account.code == FIXED_ASSET_PARENT)).scalar()
    existing = {c for (c,) in session.execute(select(Account.code)).all()}
    if code is None:
        code = next((str(n) for n in range(1230, 1290) if str(n) not in existing),
                    None)
        if code is None:
            raise ValueError(
                "Ran out of auto-numbered fixed-asset GL codes (1230-1289). Put "
                "an existing asset account code in 'asset_account_code' instead.")
    acct = Account(code=code, name=str(name)[:255], type=AccountType.ASSET,
                   parent_id=parent_id, is_postable=True)
    session.add(acct)
    session.flush()
    return acct.id


def _resolve_asset_account(session: Session, category: str,
                           override_code: Optional[str]) -> int:
    if override_code:
        a = _postable_by_code(session, override_code)
        if a is None or not a.is_postable:
            raise ValueError(
                f"Asset GL account '{override_code}' not found or not postable.")
        if a.type != AccountType.ASSET:
            raise ValueError(
                f"Asset GL account '{override_code}' is not an ASSET account.")
        return a.id

    cat = (category or "Other").strip()
    default_code = _CATEGORY_GL.get(cat.lower())
    if default_code:
        a = _postable_by_code(session, default_code)
        return a.id if a is not None else _create_asset_account(
            session, cat, code=default_code)

    # Reuse an existing asset account whose name matches the category…
    a = session.execute(
        select(Account).where(Account.type == AccountType.ASSET,
                              Account.is_postable == True,        # noqa: E712
                              Account.code.like("12%"),
                              Account.name.ilike(cat))
    ).scalars().first()
    if a is not None:
        return a.id
    # …otherwise auto-create one under Fixed Assets.
    return _create_asset_account(session, cat)


def _resolve_simple(session: Session, override_code: Optional[str],
                    default_code: str, label: str) -> int:
    code = override_code or default_code
    a = _postable_by_code(session, code)
    if a is None or not a.is_postable:
        raise ValueError(f"{label} account '{code}' not found or not postable.")
    return a.id


# --------------------------------------------------------------------------- #
# Template                                                                    #
# --------------------------------------------------------------------------- #

_HEADERS = ["code", "name", "category", "acquired_date", "cost",
            "useful_life_months", "salvage_value", "accumulated_depreciation",
            "depreciation_through", "asset_account_code",
            "accum_dep_account_code", "dep_expense_account_code"]


def template_bytes() -> bytes:
    instructions = [
        ["Trakit365 — Fixed-asset register template"], [""],
        ["One row per asset. Fill the required columns; leave the optional GL "
         "code columns blank to use sensible defaults."],
        ["This import creates the asset register only — it posts NO journal "
         "entry. Make sure each asset's cost and accumulated depreciation are "
         "already in your books (via opening balances or the original "
         "purchase); this sheet is just the detail behind those balances."],
        [""],
        ["Column", "Required?", "Notes"],
        ["code", "REQUIRED", "Unique asset code/tag, e.g. FA-001."],
        ["name", "REQUIRED", "Asset description, e.g. 'Toyota Hilux'."],
        ["category", "Optional", "Equipment, Furniture, Vehicles, Other or your "
         "own label. Blank = Other."],
        ["acquired_date", "REQUIRED", "Date acquired, YYYY-MM-DD."],
        ["cost", "REQUIRED", "Original cost (₦). Must be > 0."],
        ["useful_life_months", "REQUIRED", "Total useful life in months, e.g. 36."],
        ["salvage_value", "Optional", "Residual value at end of life (₦). "
         "Default 0. Must be less than cost."],
        ["accumulated_depreciation", "Optional", "Depreciation ALREADY booked "
         "before go-live (₦). Default 0 for a brand-new asset. Must be "
         "0..(cost − salvage)."],
        ["depreciation_through", "If accum > 0", "Month-end through which "
         "depreciation has already been booked (YYYY-MM-DD). REQUIRED whenever "
         "accumulated_depreciation is greater than 0, so future depreciation "
         "resumes from the next month instead of re-posting the past."],
        ["asset_account_code", "Optional", "GL asset account code. Blank: "
         "Equipment→1210, Furniture→1220; any other category reuses or "
         "auto-creates a 12xx account named after the category."],
        ["accum_dep_account_code", "Optional", f"Default {ACCUM_DEP_CODE} "
         "(Accumulated Depreciation)."],
        ["dep_expense_account_code", "Optional", f"Default {DEP_EXPENSE_CODE} "
         "(Depreciation Expense)."],
        [""],
        ["Examples:"],
        ["code", "name", "category", "acquired_date", "cost",
         "useful_life_months", "salvage_value", "accumulated_depreciation",
         "depreciation_through"],
        ["FA-001", "Dell servers", "Equipment", "2026-06-01", "1200000", "36",
         "0", "0", ""],
        ["FA-002", "Toyota Hilux", "Vehicles", "2024-01-15", "9000000", "60",
         "900000", "2430000", "2026-05-31"],
    ]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        pd.DataFrame(columns=_HEADERS).to_excel(xw, index=False, sheet_name=SHEET)
        pd.DataFrame(instructions).to_excel(
            xw, index=False, header=False, sheet_name="Instructions")
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        ws = xw.sheets[SHEET]
        for i, _ in enumerate(_HEADERS, start=1):
            c = ws.cell(row=1, column=i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
            ws.column_dimensions[get_column_letter(i)].width = 22
        ws.freeze_panes = "A2"
        instr = xw.sheets["Instructions"]
        for col, w in (("A", 26), ("B", 14), ("C", 72)):
            instr.column_dimensions[col].width = w
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# Import                                                                      #
# --------------------------------------------------------------------------- #

def import_assets(session: Session, df: pd.DataFrame) -> dict:
    """Create FixedAsset rows from a filled template. Per-row validation:
    bad rows are skipped with an error message; good rows are inserted. Posts
    no journal entry. Returns {created, skipped, errors}."""
    authz.require_perm("manage.assets")
    df = df.rename(columns={c: str(c).strip().lower() for c in df.columns})
    if "code" not in df.columns:
        raise ValueError("The file has no 'code' column — use the template.")

    existing = {c for (c,) in session.execute(select(FixedAsset.code)).all()}
    seen: set[str] = set()
    created = skipped = 0
    errors: list[str] = []

    for idx, row in df.iterrows():
        rno = int(idx) + 2
        code = _clean(row.get("code"))
        name = _clean(row.get("name"))
        # Wholly blank row -> ignore silently.
        if not code and not name and _num(row.get("cost")) in (0.0,) \
                and _clean(row.get("acquired_date")) is None:
            continue
        if not code:
            errors.append(f"Row {rno}: missing code."); skipped += 1; continue
        if code in existing or code in seen:
            skipped += 1; continue
        if not name:
            errors.append(f"Row {rno}: missing name."); skipped += 1; continue

        category = _clean(row.get("category")) or "Other"
        acquired = _parse_date(row.get("acquired_date"))
        if acquired is None:
            errors.append(f"Row {rno}: missing or invalid acquired_date (use YYYY-MM-DD).")
            skipped += 1; continue
        cost = _num(row.get("cost"))
        if math.isnan(cost) or cost <= 0:
            errors.append(f"Row {rno}: cost must be a number > 0."); skipped += 1; continue
        life = _int(row.get("useful_life_months"))
        if life is None or life <= 0:
            errors.append(f"Row {rno}: useful_life_months must be a whole number > 0.")
            skipped += 1; continue
        salvage = _num(row.get("salvage_value"))
        if math.isnan(salvage):
            salvage = 0.0
        if salvage < 0 or salvage >= cost:
            errors.append(f"Row {rno}: salvage_value must be between 0 and less than cost.")
            skipped += 1; continue

        accum = _num(row.get("accumulated_depreciation"))
        if math.isnan(accum):
            accum = 0.0
        depreciable = round(cost - salvage, 2)
        if accum < 0 or round(accum, 2) > depreciable:
            errors.append(f"Row {rno}: accumulated_depreciation must be between 0 "
                          f"and cost − salvage ({depreciable:,.2f}).")
            skipped += 1; continue

        last_dep: Optional[date] = None
        if round(accum, 2) > 0:
            through = _parse_date(row.get("depreciation_through"))
            if through is None:
                errors.append(f"Row {rno}: accumulated_depreciation is set, so "
                              "'depreciation_through' (the month-end already booked) "
                              "is required — otherwise depreciation would re-post the past.")
                skipped += 1; continue
            if through < acquired:
                errors.append(f"Row {rno}: depreciation_through is before acquired_date.")
                skipped += 1; continue
            last_dep = _end_of_month(through)

        try:
            # Pure lookups first; the asset account is resolved last because it
            # may auto-create — so a bad accum/dep override never leaves an
            # orphaned auto-created account behind.
            accum_acc = _resolve_simple(
                session, _clean(row.get("accum_dep_account_code")),
                ACCUM_DEP_CODE, "Accumulated depreciation")
            dep_acc = _resolve_simple(
                session, _clean(row.get("dep_expense_account_code")),
                DEP_EXPENSE_CODE, "Depreciation expense")
            asset_acc = _resolve_asset_account(
                session, category, _clean(row.get("asset_account_code")))
        except ValueError as e:
            errors.append(f"Row {rno}: {e}"); skipped += 1; continue

        session.add(FixedAsset(
            code=code, name=name, category=category, acquired_date=acquired,
            cost=round(cost, 2), salvage_value=round(salvage, 2),
            useful_life_months=life,
            depreciation_method=DepreciationMethod.STRAIGHT_LINE,
            gl_asset_account_id=asset_acc, gl_accum_dep_account_id=accum_acc,
            gl_dep_expense_account_id=dep_acc, status=AssetStatus.ACTIVE,
            accumulated_depreciation=round(accum, 2),
            last_depreciation_date=last_dep, is_active=True))
        session.flush()
        existing.add(code); seen.add(code); created += 1

    return {"created": created, "skipped": skipped, "errors": errors}
