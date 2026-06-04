"""BizClinik workbook wrapper.

Wraps a 'BizClinik Accounting and Business Software' .xlsx file and exposes
typed accessors for each module (inventory, suppliers, customers, operating
expenses, chart of accounts), plus KPI helpers and safe append-on-copy writes.

The source file is never overwritten. Mutating operations always require an
explicit output path via `save_as()`, or are auto-routed to a timestamped copy.
"""
from __future__ import annotations

import csv
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from .models import (
    Account,
    CompanyInfo,
    CustomerEntry,
    InventoryItem,
    InventoryMovement,
    OperatingEntry,
    SupplierEntry,
)


# ---- sheet/column conventions (1-based) -------------------------------------

SHEET_COMPANY = "Company Details"
SHEET_INV_LIST = "Inventory List"
SHEET_INV_MOD = "Inventory Module"
SHEET_SUP_LIST = "Supplier List"
SHEET_SUP_MOD = "Supplier Module"
SHEET_CUS_LIST = "Customer List"
SHEET_CUS_MOD = "Customer Module"
SHEET_OPS_MOD = "Operating Module"
SHEET_PNL = "Profit and Loss"
SHEET_BS = "Balance Sheet"
SHEET_QUOTE = "Sales Quotation"
SHEET_COA = "Chart of Accounts"


@dataclass(frozen=True)
class _ModuleSpec:
    sheet: str
    header_row: int
    data_start_row: int
    columns: dict  # field name -> 1-based column index


# Confirmed by inspecting Wendysrack Luxe Ltd workbook.
INVENTORY_MODULE = _ModuleSpec(
    sheet=SHEET_INV_MOD,
    header_row=8,
    data_start_row=9,
    columns={
        "code": 2,         # B
        "description": 4,  # D
        "qty_in": 5,       # E
        "qty_out": 6,      # F
        "avg_cost": 7,     # G
        "balance": 8,      # H
    },
)

SUPPLIER_MODULE = _ModuleSpec(
    sheet=SHEET_SUP_MOD,
    header_row=8,
    data_start_row=9,
    columns={
        "date": 2,
        "reference_no": 3,
        "vendor": 4,
        "code": 5,
        "description": 6,
        "qty_in": 7,
        "rate": 8,
        "total": 9,
        "vat": 10,
        "total_after_vat": 11,
        "account_name": 12,
    },
)

CUSTOMER_MODULE = _ModuleSpec(
    sheet=SHEET_CUS_MOD,
    header_row=8,
    data_start_row=9,
    columns={
        "date": 2,
        "reference_no": 3,
        "customer": 4,
        "code": 5,
        "description": 6,
        "qty_out": 7,
        "rate": 8,
        "total": 9,
        "vat": 10,
        "total_after_vat": 11,
        "account_name": 12,
    },
)

OPERATING_MODULE = _ModuleSpec(
    sheet=SHEET_OPS_MOD,
    header_row=8,
    data_start_row=9,
    columns={
        "date": 2,
        "reference_no": 3,
        "vendor": 4,
        "description": 5,
        "qty_in": 6,
        "rate": 7,
        "total": 8,
        "vat": 9,
        "total_after_vat": 10,
        "account_name": 11,
    },
)

COA = _ModuleSpec(
    sheet=SHEET_COA,
    header_row=7,
    data_start_row=8,
    columns={"code": 2, "name": 3, "category": 4},
)


# ---- core wrapper -----------------------------------------------------------


class BizClinikWorkbook:
    """Read/write wrapper around a BizClinik accounting .xlsx file.

    Use `read_only=True` for fast inspection. For appends/writes, use the
    default (read_only=False) and call `save_as()` to persist to a new file.
    """

    def __init__(self, path: str | Path, read_only: bool = True):
        self.path = Path(path)
        if not self.path.exists():
            raise FileNotFoundError(self.path)
        # data_only=True reads cached values from the last Excel save.
        # For mutation we re-open without data_only so formulas survive.
        self._wb = load_workbook(self.path, data_only=read_only, read_only=read_only)
        self._read_only = read_only

    # ---- basic access -------------------------------------------------------

    @property
    def workbook(self) -> Workbook:
        return self._wb

    @property
    def sheet_names(self) -> list[str]:
        return list(self._wb.sheetnames)

    def _sheet(self, name: str):
        if name not in self._wb.sheetnames:
            raise KeyError(f"Sheet not found: {name}. Available: {self.sheet_names}")
        return self._wb[name]

    # ---- company info -------------------------------------------------------

    def company(self) -> CompanyInfo:
        ws = self._sheet(SHEET_COMPANY)
        # Form-style: label in col A, value in col B. Scan rows 1..max.
        info = CompanyInfo()
        label_map = {
            "company name": "name",
            "rc number": "rc_number",
            "registered address": "address",
            "address": "address",
            "vat no": "vat_no",
            "vat number": "vat_no",
            "email": "email",
            "phone": "phone",
            "telephone": "phone",
        }
        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            value = row[1] if len(row) > 1 else None
            if label in label_map and value not in (None, ""):
                setattr(info, label_map[label], value)
        return info

    # ---- inventory ----------------------------------------------------------

    def inventory_items(self) -> list[InventoryItem]:
        """Read the master inventory list (codes + descriptions)."""
        ws = self._sheet(SHEET_INV_LIST)
        items: list[InventoryItem] = []
        header_row = self._find_header_row(ws, needles=("product code", "code"))
        if header_row is None:
            return items
        code_col, desc_col = self._locate_inv_list_cols(ws, header_row)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            code = row[code_col - 1] if code_col else None
            desc = row[desc_col - 1] if desc_col else None
            if code in (None, "") and desc in (None, ""):
                continue
            items.append(InventoryItem(code=str(code) if code is not None else "",
                                       description=str(desc) if desc is not None else ""))
        return items

    def inventory_movements(self) -> list[InventoryMovement]:
        return list(self._iter_records(INVENTORY_MODULE, InventoryMovement))

    def stock_balance(self) -> dict[str, float]:
        """Net qty per product code from the Inventory Module."""
        bal: dict[str, float] = {}
        for m in self.inventory_movements():
            if not m.code:
                continue
            bal[m.code] = bal.get(m.code, 0.0) + (m.qty_in or 0) - (m.qty_out or 0)
        return bal

    # ---- suppliers / customers / opex --------------------------------------

    def supplier_entries(self) -> list[SupplierEntry]:
        return list(self._iter_records(SUPPLIER_MODULE, SupplierEntry))

    def customer_entries(self) -> list[CustomerEntry]:
        return list(self._iter_records(CUSTOMER_MODULE, CustomerEntry))

    def operating_entries(self) -> list[OperatingEntry]:
        return list(self._iter_records(OPERATING_MODULE, OperatingEntry))

    # ---- chart of accounts --------------------------------------------------

    def chart_of_accounts(self) -> list[Account]:
        ws = self._sheet(SHEET_COA)
        rows: list[Account] = []
        for r in ws.iter_rows(min_row=COA.data_start_row, values_only=True):
            code = r[COA.columns["code"] - 1]
            name = r[COA.columns["name"] - 1]
            cat = r[COA.columns["category"] - 1]
            if name in (None, ""):
                continue
            rows.append(Account(code=code, name=str(name), category=cat))
        return rows

    # ---- KPIs ---------------------------------------------------------------

    def kpis(self) -> dict:
        sup = self.supplier_entries()
        cus = self.customer_entries()
        ops = self.operating_entries()
        inv = self.inventory_movements()

        def _num(x):
            try:
                return float(x) if x is not None else 0.0
            except (TypeError, ValueError):
                return 0.0

        total_purchases = sum(_num(e.total) for e in sup)
        total_sales = sum(_num(e.total) for e in cus)
        total_opex = sum(_num(e.total) for e in ops)
        sales_vat = sum(_num(e.vat) for e in cus)
        purchase_vat = sum(_num(e.vat) for e in sup)

        gross_profit = total_sales - total_purchases
        net_profit = gross_profit - total_opex

        return {
            "purchases_total": total_purchases,
            "sales_total": total_sales,
            "operating_expenses_total": total_opex,
            "gross_profit": gross_profit,
            "net_profit": net_profit,
            "vat_on_sales": sales_vat,
            "vat_on_purchases": purchase_vat,
            "vat_net_payable": sales_vat - purchase_vat,
            "inventory_rows": len(inv),
            "supplier_rows": len(sup),
            "customer_rows": len(cus),
            "operating_rows": len(ops),
        }

    # ---- safe writes --------------------------------------------------------

    def append_supplier(self, entry: SupplierEntry) -> None:
        self._append(SUPPLIER_MODULE, entry.__dict__)

    def append_customer(self, entry: CustomerEntry) -> None:
        self._append(CUSTOMER_MODULE, entry.__dict__)

    def append_operating(self, entry: OperatingEntry) -> None:
        self._append(OPERATING_MODULE, entry.__dict__)

    def append_inventory_movement(self, movement: InventoryMovement) -> None:
        self._append(INVENTORY_MODULE, movement.__dict__)

    def save_as(self, out_path: str | Path) -> Path:
        """Save the (possibly mutated) workbook to a NEW path.

        Honors the user's file-safety rule: never overwrite existing files.
        Raises FileExistsError if out_path already exists.
        """
        out = Path(out_path)
        if out.exists():
            raise FileExistsError(
                f"Refusing to overwrite existing file: {out}. "
                "Pass a new path or use save_timestamped()."
            )
        out.parent.mkdir(parents=True, exist_ok=True)
        if self._read_only:
            raise RuntimeError(
                "Workbook opened read_only=True; reopen with read_only=False to write."
            )
        self._wb.save(out)
        return out

    def save_timestamped(self, out_dir: Optional[str | Path] = None,
                         suffix: str = "updated") -> Path:
        """Save to '<stem>__<suffix>__<YYYYMMDD-HHMMSS>.xlsx' next to source."""
        out_dir = Path(out_dir) if out_dir else self.path.parent
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        out = out_dir / f"{self.path.stem}__{suffix}__{ts}.xlsx"
        return self.save_as(out)

    @classmethod
    def copy_for_editing(cls, source: str | Path,
                         out_dir: Optional[str | Path] = None) -> "BizClinikWorkbook":
        """Make a timestamped copy of `source` and open it for editing.

        The source file is never touched. Returns a writable wrapper around
        the copy.
        """
        src = Path(source)
        out_dir = Path(out_dir) if out_dir else src.parent
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        dst = out_dir / f"{src.stem}__editing__{ts}.xlsx"
        if dst.exists():
            raise FileExistsError(dst)
        out_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(src, dst)
        return cls(dst, read_only=False)

    # ---- exporters ----------------------------------------------------------

    def export_csv(self, module: str, out_path: str | Path) -> Path:
        """Export a module's records to CSV.

        module in {'inventory','suppliers','customers','operating','coa'}.
        """
        out = Path(out_path)
        if out.exists():
            raise FileExistsError(out)
        rows = self._records_for(module)
        out.parent.mkdir(parents=True, exist_ok=True)
        if not rows:
            out.write_text("", encoding="utf-8")
            return out
        fieldnames = list(rows[0].keys())
        with out.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerows(rows)
        return out

    def export_json(self, module: str, out_path: str | Path) -> Path:
        out = Path(out_path)
        if out.exists():
            raise FileExistsError(out)
        rows = self._records_for(module)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")
        return out

    # ---- internals ----------------------------------------------------------

    def _records_for(self, module: str) -> list[dict]:
        m = module.lower()
        if m in ("inventory", "inv", "stock"):
            return [r.to_dict() for r in self.inventory_movements()]
        if m in ("suppliers", "supplier", "purchases"):
            return [r.to_dict() for r in self.supplier_entries()]
        if m in ("customers", "customer", "sales"):
            return [r.to_dict() for r in self.customer_entries()]
        if m in ("operating", "ops", "opex", "expenses"):
            return [r.to_dict() for r in self.operating_entries()]
        if m in ("coa", "accounts", "chart_of_accounts"):
            return [r.to_dict() for r in self.chart_of_accounts()]
        if m in ("company",):
            return [self.company().to_dict()]
        raise ValueError(f"Unknown module: {module}")

    # Fields that, if all empty/#N/A/0, mean the row is just template noise.
    _SIGNAL_FIELDS = {
        "date", "vendor", "customer", "code", "qty_in", "qty_out",
        "rate", "total", "total_after_vat",
    }
    # Several module sheets have phantom max_row=1,048,576. Stop after this
    # many consecutive empty rows to avoid streaming through the full sheet.
    _EMPTY_ROW_RUN_LIMIT = 50

    def _iter_records(self, spec: _ModuleSpec, cls) -> Iterable:
        # Cache per-module result so re-readers (financials calls
        # supplier/inventory more than once) don't re-stream the sheet.
        cache_key = (spec.sheet, cls.__name__)
        if not hasattr(self, "_record_cache"):
            self._record_cache = {}
        if cache_key in self._record_cache:
            yield from self._record_cache[cache_key]
            return

        ws = self._sheet(spec.sheet)
        cols = spec.columns
        max_col = max(cols.values())
        signal_keys = self._SIGNAL_FIELDS & cols.keys()

        def _empty(v):
            return v in (None, "", 0) or (isinstance(v, str) and v.strip() == "#N/A")

        records = []
        empty_run = 0
        for row in ws.iter_rows(min_row=spec.data_start_row,
                                max_col=max_col, values_only=True):
            kwargs = {f: row[idx - 1] if len(row) >= idx else None
                      for f, idx in cols.items()}
            if all(_empty(kwargs.get(k)) for k in signal_keys):
                empty_run += 1
                if empty_run >= self._EMPTY_ROW_RUN_LIMIT:
                    break
                continue
            empty_run = 0
            records.append(cls(**kwargs))

        self._record_cache[cache_key] = records
        yield from records

    def _append(self, spec: _ModuleSpec, fields: dict) -> None:
        if self._read_only:
            raise RuntimeError("Workbook is read-only; reopen with read_only=False.")
        ws = self._sheet(spec.sheet)
        target_row = self._first_empty_row(ws, spec)
        for fname, col in spec.columns.items():
            if fname in fields and fields[fname] is not None:
                ws.cell(row=target_row, column=col, value=fields[fname])

    def _first_empty_row(self, ws, spec: _ModuleSpec) -> int:
        """Find the first row at/after data_start_row with no key column set."""
        key_col = next(iter(spec.columns.values()))  # first tracked column
        row = spec.data_start_row
        # Cap the scan to avoid the 1,048,576 phantom max_row some sheets have.
        for r in range(spec.data_start_row, spec.data_start_row + 100_000):
            if ws.cell(row=r, column=key_col).value in (None, ""):
                return r
            row = r
        return row + 1

    @staticmethod
    def _find_header_row(ws, needles: tuple[str, ...]) -> Optional[int]:
        needles_l = [n.lower() for n in needles]
        for r_idx, row in enumerate(ws.iter_rows(values_only=True, max_row=30), start=1):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).strip().lower()
                if any(n in s for n in needles_l):
                    return r_idx
        return None

    @staticmethod
    def _locate_inv_list_cols(ws, header_row: int) -> tuple[Optional[int], Optional[int]]:
        code_col = desc_col = None
        for c_idx, val in enumerate(
            next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)),
            start=1,
        ):
            if val is None:
                continue
            s = str(val).strip().lower()
            if "code" in s and code_col is None:
                code_col = c_idx
            elif "desc" in s and desc_col is None:
                desc_col = c_idx
        return code_col, desc_col
