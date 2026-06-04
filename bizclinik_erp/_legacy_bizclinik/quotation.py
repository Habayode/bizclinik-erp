"""Sales Quotation generator.

Produces a standalone quotation .xlsx that mirrors the BizClinik 'Sales
Quotation' sheet layout. Decoupled from any specific source workbook so it
can be generated from the wrapper API or the Streamlit UI.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import CompanyInfo


# ---- data shapes -----------------------------------------------------------


@dataclass
class QuotationParty:
    name: str
    address: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


@dataclass
class QuotationLine:
    code: Optional[str]
    description: str
    qty: float
    rate: float
    vat_rate: float = 0.0  # e.g. 0.075 for 7.5%

    @property
    def subtotal(self) -> float:
        return float(self.qty) * float(self.rate)

    @property
    def vat_amount(self) -> float:
        return self.subtotal * float(self.vat_rate)

    @property
    def total(self) -> float:
        return self.subtotal + self.vat_amount


@dataclass
class Quotation:
    company: CompanyInfo
    customer: QuotationParty
    ref_no: str
    issue_date: date = field(default_factory=lambda: datetime.now().date())
    valid_until: Optional[date] = None
    lines: list[QuotationLine] = field(default_factory=list)
    notes: Optional[str] = None
    currency: str = "₦"

    @property
    def subtotal(self) -> float:
        return sum(l.subtotal for l in self.lines)

    @property
    def vat_total(self) -> float:
        return sum(l.vat_amount for l in self.lines)

    @property
    def grand_total(self) -> float:
        return self.subtotal + self.vat_total

    def amount_in_words(self) -> str:
        return _amount_in_words(self.grand_total, self.currency)


# ---- writer ----------------------------------------------------------------


_THIN = Side(style="thin", color="999999")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FILL = PatternFill("solid", start_color="1F3864")
_HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_LABEL_FONT = Font(name="Arial", bold=True, size=10)
_BODY_FONT = Font(name="Arial", size=10)
_TITLE_FONT = Font(name="Arial", bold=True, size=18, color="1F3864")


def write_quotation_xlsx(quote: Quotation, out_path: str | Path) -> Path:
    """Write the quotation to a new .xlsx file. Refuses to overwrite."""
    out = Path(out_path)
    if out.exists():
        raise FileExistsError(
            f"Refusing to overwrite: {out}. Choose a new filename."
        )
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Quotation"

    # Column widths matching BizClinik layout (12 cols).
    widths = [3, 6, 14, 30, 6, 6, 8, 14, 12, 16, 3, 3]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("B1:J1")
    ws["B1"] = "SALES QUOTATION"
    ws["B1"].font = _TITLE_FONT
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Company info (left) and customer info (right)
    ws["B4"] = "Company"
    ws["B4"].font = _LABEL_FONT
    ws["B5"] = quote.company.name or ""
    ws["B6"] = quote.company.address or ""
    ws["B7"] = quote.company.email or ""
    ws["B8"] = quote.company.phone or ""
    ws["B9"] = quote.company.rc_number or ""
    for r in range(5, 10):
        ws.cell(row=r, column=2).font = _BODY_FONT

    ws["H4"] = "Bill To"
    ws["H4"].font = _LABEL_FONT
    ws["H5"] = quote.customer.name
    ws["H6"] = quote.customer.address or ""
    ws["H7"] = quote.customer.email or ""
    ws["H8"] = quote.customer.phone or ""
    for r in range(5, 10):
        ws.cell(row=r, column=8).font = _BODY_FONT

    # Reference info
    ws["B11"] = "Ref No"
    ws["C11"] = quote.ref_no
    ws["B12"] = "Date"
    ws["C12"] = quote.issue_date.isoformat()
    if quote.valid_until:
        ws["B13"] = "Valid Until"
        ws["C13"] = quote.valid_until.isoformat()
    for r in (11, 12, 13):
        ws.cell(row=r, column=2).font = _LABEL_FONT
        ws.cell(row=r, column=3).font = _BODY_FONT

    # Line items header
    header_row = 15
    headers = ["S/N", "Code", "Description", "Qty", "Rate",
               "Subtotal", "VAT %", "VAT Amt", "Total"]
    # Column mapping: B=S/N, C=Code, D-F=Description (merged), G=Qty,
    # H=Rate, I=Subtotal, J=VAT%, K=VAT Amt, L=Total
    layout_cols = [2, 3, 4, 7, 8, 9, 10, 11, 12]
    for c, h in zip(layout_cols, headers):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.merge_cells(start_row=header_row, start_column=4,
                   end_row=header_row, end_column=6)
    ws.row_dimensions[header_row].height = 20

    # Line items
    first_data_row = header_row + 1
    for i, line in enumerate(quote.lines, start=1):
        r = header_row + i
        ws.cell(row=r, column=2, value=i)
        ws.cell(row=r, column=3, value=line.code or "")
        ws.cell(row=r, column=4, value=line.description)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=7, value=line.qty)
        ws.cell(row=r, column=8, value=line.rate)
        # Excel formulas so downstream edits recalc properly.
        ws.cell(row=r, column=9, value=f"=G{r}*H{r}")
        ws.cell(row=r, column=10, value=line.vat_rate)
        ws.cell(row=r, column=10).number_format = "0.00%"
        ws.cell(row=r, column=11, value=f"=I{r}*J{r}")
        ws.cell(row=r, column=12, value=f"=I{r}+K{r}")
        for c in layout_cols:
            cell = ws.cell(row=r, column=c)
            cell.font = _BODY_FONT
            cell.border = _BORDER
            cell.alignment = Alignment(vertical="center")
        # Currency formats
        for c in (8, 9, 11, 12):
            ws.cell(row=r, column=c).number_format = '#,##0.00;(#,##0.00);-'

    last_data_row = header_row + len(quote.lines)

    # Totals block
    totals_row = last_data_row + 2
    ws.cell(row=totals_row, column=8, value="Subtotal").font = _LABEL_FONT
    ws.cell(row=totals_row, column=9,
            value=f"=SUM(I{first_data_row}:I{last_data_row})")
    ws.cell(row=totals_row + 1, column=8, value="VAT").font = _LABEL_FONT
    ws.cell(row=totals_row + 1, column=11,
            value=f"=SUM(K{first_data_row}:K{last_data_row})")
    ws.cell(row=totals_row + 2, column=8, value="Grand Total").font = Font(
        name="Arial", bold=True, size=12)
    ws.cell(row=totals_row + 2, column=12,
            value=f"=I{totals_row}+K{totals_row + 1}")
    for c, r in [(9, totals_row), (11, totals_row + 1), (12, totals_row + 2)]:
        cell = ws.cell(row=r, column=c)
        cell.number_format = '#,##0.00;(#,##0.00);-'
        cell.font = _LABEL_FONT
        cell.border = _BORDER

    # Amount in words
    words_row = totals_row + 4
    ws.cell(row=words_row, column=2, value="Amount in words").font = _LABEL_FONT
    ws.merge_cells(start_row=words_row, start_column=3,
                   end_row=words_row, end_column=12)
    ws.cell(row=words_row, column=3, value=quote.amount_in_words()).font = _BODY_FONT

    if quote.notes:
        notes_row = words_row + 2
        ws.cell(row=notes_row, column=2, value="Notes").font = _LABEL_FONT
        ws.merge_cells(start_row=notes_row, start_column=3,
                       end_row=notes_row + 2, end_column=12)
        ws.cell(row=notes_row, column=3, value=quote.notes).alignment = Alignment(
            wrap_text=True, vertical="top")

    # Signatures
    sig_row = words_row + 6
    ws.cell(row=sig_row, column=2, value="Company Signature").font = _LABEL_FONT
    ws.cell(row=sig_row, column=8, value="Customer Signature").font = _LABEL_FONT

    wb.save(out)
    return out


# ---- amount in words -------------------------------------------------------


_ONES = [
    "", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
    "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen",
    "seventeen", "eighteen", "nineteen",
]
_TENS = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy",
         "eighty", "ninety"]
_SCALES = [("trillion", 10**12), ("billion", 10**9), ("million", 10**6),
           ("thousand", 10**3)]


def _under_thousand(n: int) -> str:
    if n == 0:
        return ""
    if n < 20:
        return _ONES[n]
    if n < 100:
        rest = n % 10
        return _TENS[n // 10] + (f"-{_ONES[rest]}" if rest else "")
    rest = n % 100
    head = f"{_ONES[n // 100]} hundred"
    return f"{head} and {_under_thousand(rest)}" if rest else head


def _int_to_words(n: int) -> str:
    if n == 0:
        return "zero"
    parts: list[str] = []
    for word, scale in _SCALES:
        if n >= scale:
            chunk = n // scale
            parts.append(f"{_under_thousand(chunk)} {word}")
            n %= scale
    if n:
        parts.append(_under_thousand(n))
    return " ".join(parts).strip()


def _amount_in_words(amount: float, currency_symbol: str = "₦") -> str:
    # Strip symbol for word output; symbol stays in the printed totals.
    currency_word = {"₦": "naira", "$": "dollars", "€": "euros",
                     "£": "pounds"}.get(currency_symbol, "")
    sign = "minus " if amount < 0 else ""
    amount = abs(amount)
    whole = int(amount)
    fraction = round((amount - whole) * 100)
    main = _int_to_words(whole)
    out = f"{sign}{main}"
    if currency_word:
        out += f" {currency_word}"
    if fraction:
        out += f" and {_int_to_words(fraction)} kobo" if currency_word == "naira" \
               else f" and {_int_to_words(fraction)} / 100"
    return out.strip().capitalize() + " only"
